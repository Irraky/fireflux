import pprint
from enum import Enum
from pydantic import BaseModel
import json
import csv
import ipaddress
from io import TextIOWrapper
from typing import Iterable, Tuple
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from tabulate import tabulate


# --- https://docs.opnsense.org/manual/firewall.html#the-basics


class Action(str, Enum):
    """Firewall actions"""

    Pass = "pass"
    Block = "block"
    Reject = "reject"


class IpVer(str, Enum):
    """Internet Protocol verions"""

    V4 = "IPv4"
    V6 = "IPv6"
    Both = "IPv4+IPv6"


class Protocol(str, Enum):
    """Network protocols"""

    TCP = "TCP"
    UDP = "UDP"
    TPC_UDP = "TCP/UDP"
    # I added all protocols supported by pfSense but I am unsure if we can/want to support all of them
    ICMP = "ICMP"
    ESP = "ESP"
    AH = "AH"
    GRE = "GRE"
    EoIP = "EoIP"
    IPV6 = "IPV6"
    IGMP = "IGMP"
    PIM = "PIM"
    OSPF = "OSPF"
    SCTP = "SCTP"
    CARP = "CARP"
    PFSYNC = "PFSYNC"
    ETHERIP = "ETHERIP"


class NetworkFilter(BaseModel):
    """Network filter"""

    inverted: bool
    network: ipaddress.IPv4Network | ipaddress.IPv6Network | str | None

    def to_str(self) -> str:
        inverted = "!" if self.inverted else ""
        network = self.network if self.network is not None else "*"
        return f"{inverted}{network}"

    def to_str_2(self) -> str:
        network = self.network if self.network is not None else "*"
        return f"{network}"

    @staticmethod
    def from_str(value):
        inverted = False
        if value.startswith("!"):
            inverted = True
            value = value[1:]
        if value == "*":
            network = None
        else:
            try:
                network = ipaddress.ip_network(value)
            except:
                network = value
        return NetworkFilter(inverted=inverted, network=network)


class PortRange(BaseModel):
    """Port range filter"""

    range: int | Tuple[int, int] | None

    def to_str(self) -> str:
        if self.range is None:
            return "*"
        elif type(self.range) is tuple:
            return f"{self.range[0]}-{self.range[1]}"
        else:
            return f"{self.range}"

    @staticmethod
    def from_str(value):
        if value == "*" or value == "":
            range = None
        else:
            range = [int(nb) for nb in value.split("-", 1)]
            if len(range) == 1:
                range = range[0]
            else:
                range = (range[0], range[1])

        return PortRange(range=range)


class Rule(BaseModel):
    """Generic firewall rule"""

    description: str | None
    action: Action
    interface: str  # TODO interface type
    ip_ver: IpVer
    protocol: Protocol | None
    source: NetworkFilter
    source_ports: PortRange
    destination: NetworkFilter
    destination_port: PortRange

    def get_array(self):
        return [
            self.description,
            self.action.value,
            self.interface,
            self.ip_ver.value,
            self.protocol.value if self.protocol != None else "*",
            self.source.to_str_2(),
            self.test_source_ports(),
            self.destination.to_str_2(),
            self.test_destination_port(),
        ]

    def get_direction(self):
        return self.source.to_str_2(), self.destination.to_str_2()

    def test_source_ports(self):
        if self.source_ports.range is None:
            return "*"
        else:
            return self.source_ports.range

    def test_destination_port(self):
        if self.destination_port.range is None:
            return "*"
        else:
            return self.destination_port.range


# --- Rules serialization


def __ugly_hack(dict):
    dict.source = dict.source.to_str()
    dict.destination = dict.destination.to_str()
    dict.source_ports = dict.source_ports.to_str()
    dict.destination_port = dict.destination_port.to_str()
    return dict.dict()


def __ugly_hack2(dict):
    dict["source"] = NetworkFilter.from_str(dict["source"])
    dict["destination"] = NetworkFilter.from_str(dict["destination"])
    dict["source_ports"] = PortRange.from_str(dict["source_ports"])
    dict["destination_port"] = PortRange.from_str(dict["destination_port"])
    if dict["protocol"] == "":
        dict["protocol"] = None
    return dict


def rules_to_json(dst: TextIOWrapper, rules: Iterable[Rule]):
    """Serialize rules into a JSON stream"""
    json.dump([__ugly_hack(r) for r in rules], dst)


def rules_from_json(src: TextIOWrapper) -> list[Rule]:
    """Deserialize rules from a JSON stream"""
    return [Rule.parse_obj(__ugly_hack2(dict)) for dict in json.load(src)]


def rules_to_csv(dst: TextIOWrapper, rules: Iterable[Rule]):
    """Serialize rules into a CSV stream"""
    fieldnames = list(Rule.schema()["properties"].keys())
    w = csv.DictWriter(dst, fieldnames=fieldnames)
    w.writeheader()
    for r in rules:
        w.writerow(__ugly_hack(r))


def rules_from_csv(src: TextIOWrapper) -> list[Rule]:
    """Deserialize rules from a CSV stream"""
    return [Rule.parse_obj(__ugly_hack2(dict)) for dict in csv.DictReader(src)]


def auto_adjust_cell_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter

        for cell in col:
            cell_lines = str(cell.value).split("\n")
            max_line_length = max([len(line) for line in cell_lines])
            if max_line_length > max_length:
                max_length = max_line_length

        adjusted_width = max_length
        worksheet.column_dimensions[column].width = adjusted_width


def auto_adjust_cell_height(worksheet, wrap_text=True, default_height=15):
    for row in worksheet:
        max_lines = 1
        for cell in row:
            if wrap_text:
                cell.alignment = openpyxl.styles.Alignment(
                    wrapText=True, vertical="center"
                )

            cell_lines = str(cell.value).count("\n") + 1
            if cell_lines > max_lines:
                max_lines = cell_lines

        adjusted_height = max_lines * default_height
        worksheet.row_dimensions[row[0].row].height = adjusted_height


def worksheet_stylization(worksheet):
    for j in range(0, worksheet.max_column):
        for i in range(1, worksheet.max_row + 1):
            if j == 0 or i == 1:
                cell = worksheet[i][j]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    wrapText=True, horizontal="center", vertical="center"
                )


def get_dict(rules: Iterable[Rule]):
    rules_dict = {}
    for r in rules:
        current_rule = r.get_array()
        direction = r.get_direction()
        protocol = r.protocol.value if r.protocol != None else "*"
        if direction in rules_dict:
            if protocol in rules_dict[direction]:
                rules_dict[direction][protocol].append(current_rule)
            else:
                rules_dict[direction] = {protocol: [current_rule]}
        else:
            rules_dict[direction] = {protocol: [current_rule]}

    return rules_dict


def rules_to_excel(path: str, rules: list[Rule]):
    rules_dict = get_dict(rules)
    sources = []
    destinations = []
    for key in rules_dict.keys():
        if key[0] not in sources:
            sources.append(key[0])
        if key[1] not in destinations:
            destinations.append(key[1])

    flow_matrix = np.empty((len(sources) + 1, len(destinations) + 1), dtype=object)

    flow_matrix[0][0] = "Source / Destination"

    for y in range(len(flow_matrix)):
        for x in range(len(flow_matrix[y])):
            if y == 0 and x == 0:
                pass
            elif y == 0 and x > 0:
                flow_matrix[y][x] = destinations[x - 1]
            elif y > 0 and x == 0:
                flow_matrix[y][x] = sources[y - 1]
            elif (sources[y - 1], destinations[x - 1]) in rules_dict:
                rules = rules_dict[(sources[y - 1], destinations[x - 1])]
                res = ""
                for protocol in rules.keys():
                    for rule in rules[protocol]:
                        res += f"Protocol : {rule[4]} - Port : {rule[8]}\n"
                    flow_matrix[y][x] = res[:-1]

    data_frame = pd.DataFrame(flow_matrix)

    writer = pd.ExcelWriter(path)
    data_frame.to_excel(
        writer, sheet_name="Flow matrix", header=False, index=False, na_rep=""
    )
    writer.close()

    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active

    auto_adjust_cell_width(worksheet)
    auto_adjust_cell_height(worksheet)
    worksheet_stylization(worksheet)

    output = "./resources/output.xlsx"
    workbook.save(output)


def visualize(path, espace=2):
    df = pd.read_excel(path, sheet_name="Flow matrix", engine="openpyxl")
    df.fillna("", inplace=True)

    headers = df.columns
    table = tabulate(
        df,
        headers,
        showindex=False,
        tablefmt="psql",
        colalign=("center",) * len(headers),
        numalign="left",
        stralign="left",
    )

    print("\n")
    print(table)
    print("\n")
